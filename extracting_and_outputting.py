import time
import os
import sys
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os.path
from os import path


#class Sensor, which we can use to access the object's(sensor's) data
class Sensor:
    def __init__ (self, sensor_number, column, row):
        self.sensor_number = sensor_number
        self.COLUMN = column
        self.row = row
        self.DATE_COLUMN = column - 1
        self.date_log = ''
        self.LOG_FILE = ''
        self.current_date_log = ''
        self.log_data = ''


#function for saving the excel file
def saveWb():
    wb.save(XLSX_FILE_NAME)
    print("Saved")


def reloadWorkbook():
    global wb
    global sheet
    try:
        wb = openpyxl.load_workbook(PATH)
        print("Reloading")
    except Exception as e:
        print("Could not find workbook")
    sheet = wb.active


#function for writing the data to the spreadsheet
def writeData(sensor):
    if(sensor.log_data != '' and sensor.log_data != None):
        sensor.row = sensor.row + 1
        sheet.cell(row = sensor.row, column = sensor.COLUMN - 1).value = sensor.date_log
        sheet.cell(row = sensor.row, column = sensor.COLUMN).value = int(sensor.log_data[:4], 16)
        print("Writing data {} in row {}".format(sensor.log_data, sensor.row))       
            
#function for checking and outputting the data
def output(sensor):
    while (not os.path.exists(sensor.LOG_FILE)):
        print("Could not find file LOG" + str(sensor.sensor_number + 1))
        time.sleep(10)

    sensor.curr_date_log = time.ctime(os.stat(sensor.LOG_FILE)[8])
    
    if (sensor.date_log != sensor.curr_date_log):
        sensor.log_data = open(sensor.LOG_FILE, "r").read()
        sensor.date_log = sensor.curr_date_log
        writeData(sensor)


#Function to find the last row that has a value inside
def findLastRow(sensor):
    i = 2
    for col in sheet.iter_cols(min_row = 2, min_col = sensor.COLUMN, max_col = sensor.COLUMN):
        for cell in col:
            if cell.value != None:
                sensor.row = i
                print(cell.value)
                i = i + 1

        
#initializing the sensors by sending them to the class Sensor
sensor0 = Sensor(0, 3, 1)
sensor1 = Sensor(1, 5, 1)
sensor2 = Sensor(2, 7, 1)


#declaring the path to the cx-programmer logs
sensor0.LOG_FILE = "C:\\Users\\Ico\\AppData\\Local\\Temp\\SimulatorData\\CARD\\LOG1.txt"
sensor1.LOG_FILE = "C:\\Users\\Ico\\AppData\\Local\\Temp\\SimulatorData\\CARD\\LOG2.txt"
sensor2.LOG_FILE = "C:\\Users\\Ico\\AppData\\Local\\Temp\\SimulatorData\\CARD\\LOG3.txt"


#set the excel file name and initialize a reference to a workbook and a sheet
XLSX_FILE_NAME = "sensor_data.xlsx"
PATH = "C:\\Users\\Ico\\Desktop\\Dev\\Python\\Thesis\\" + XLSX_FILE_NAME
wb = Workbook()
reloadWorkbook()
sheet = wb.active


#Finding the row of the last empty cell of the sensor's column in the excel spreadsheet
findLastRow(sensor0)
findLastRow(sensor1)
findLastRow(sensor2)


if (__name__ == '__main__'):

#main loop of the system
    while(True):
        reloadWorkbook()
        
        output(sensor0)
        output(sensor1)
        output(sensor2)
        
        saveWb()
        
        time.sleep(10)
        


